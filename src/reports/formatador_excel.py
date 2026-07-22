import os
import openpyxl

from copy import copy

from openpyxl import load_workbook

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.chart import (
    PieChart,
    Reference,
    BarChart
)

from openpyxl.chart.label import DataLabelList

from openpyxl.formatting.rule import DataBarRule

from openpyxl.utils import get_column_letter


AZUL_ESCURO = "1F4E78"
AZUL_CLARO = "D9EAF7"
VERDE = "E2F0D9"
VERMELHO = "FCE4D6"
AMARELO = "FFF2CC"
FONTE_PADRAO = "Calibri"



# ==========================================
# AJUSTAR COLUNAS
# ==========================================


def ajustar_largura_colunas(planilha):

    for coluna in planilha.columns:

        maior = 0

        letra = get_column_letter(
            coluna[0].column
        )


        for celula in coluna:

            if celula.value:

                tamanho = len(
                    str(celula.value)
                )

                if tamanho > maior:
                    maior = tamanho


        planilha.column_dimensions[
            letra
        ].width = maior + 3




# ==========================================
# CABEÇALHO
# ==========================================


def estilizar_cabecalho(planilha):

    preenchimento = PatternFill(
        "solid",
        fgColor=AZUL_CLARO
    )


    for celula in planilha[1]:

        celula.font = Font(
            bold=True,
            color=AZUL_ESCURO,
            name=FONTE_PADRAO
        )

        celula.fill = preenchimento

        celula.alignment = Alignment(
            horizontal="center"
        )




# ==========================================
# FILTROS
# ==========================================


def aplicar_filtro(planilha):

    planilha.auto_filter.ref = (
        planilha.dimensions
    )



def congelar_cabecalho(planilha):

    planilha.freeze_panes = "A2"





# ==========================================
# FORMATAR PLANILHAS
# ==========================================


def formatar_planilhas(caminho):

    wb = load_workbook(
        caminho
    )


    for nome in wb.sheetnames:

        ws = wb[nome]


        estilizar_cabecalho(
            ws
        )


        ajustar_largura_colunas(
            ws
        )


        aplicar_filtro(
            ws
        )


        congelar_cabecalho(
            ws
        )


    wb.save(
        caminho
    )





# ==========================================
# CRIAR DASHBOARD
# ==========================================


def criar_dashboard(caminho):

    wb = load_workbook(caminho)


    if "Dashboard IA" in wb.sheetnames:
        del wb["Dashboard IA"]


    dashboard = wb.create_sheet(
        "Dashboard IA",
        0
    )


    # ==============================
    # TÍTULO
    # ==============================

    dashboard["A1"] = "RELATÓRIO DE DESEMPENHO DA IA"

    dashboard.merge_cells(
        "A1:J1"
    )

    dashboard["A1"].alignment = Alignment(
        horizontal="center"
    )

    dashboard["A1"].font = Font(
        bold=True,
        size=16,
        color="FFFFFF"
    )
    dashboard["A1"].fill = PatternFill("solid", fgColor=AZUL_ESCURO)


    # ==============================
    # TAMANHO DAS COLUNAS
    # ==============================

    larguras = {

        "A":20,
        "B":15,
        "C":5,

        "D":25,
        "E":15,
        "F":5,

        "G":20,
        "H":15,
        "I":5,

        "J":15,
        "K":15

    }


    for coluna, largura in larguras.items():

        dashboard.column_dimensions[coluna].width = largura



    # ==============================
    # CARDS KPI
    # ==============================


    cards = [

        (
            "A3",
            "A4",
            "TOTAL ANALISADO",
            "=COUNTA('Dados Analisados'!A:A)-1"
        ),


        (
            "D3",
            "D4",
            "PROBLEMAS CONFIRMADOS",
            '=COUNTIF(\'Dados Analisados\'!G:G,"Problema confirmado")'
        ),


        (
            "G3",
            "G4",
            "FALSOS POSITIVOS",
            '=COUNTIF(\'Dados Analisados\'!G:G,"Falso positivo")'
        ),


        (
            "J3",
            "J4",
            "PENDENTES",
            '=COUNTIF(\'Dados Analisados\'!G:G,"Pendente")'
        )

    ]



    for titulo, valor, nome, formula in cards:


        dashboard.merge_cells(
            f"{titulo}:{chr(ord(titulo[0])+1)}3"
        )


        dashboard.merge_cells(
            f"{valor}:{chr(ord(valor[0])+1)}4"
        )


        dashboard[titulo] = nome


        dashboard[titulo].font = Font(
            bold=True,
            size=11
        )


        dashboard[titulo].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        dashboard[titulo].fill = PatternFill(
            "solid",
            fgColor={
                "TOTAL ANALISADO": AZUL_CLARO,
                "PROBLEMAS CONFIRMADOS": VERDE,
                "FALSOS POSITIVOS": VERMELHO,
                "PENDENTES": AMARELO
            }[nome]
        )



        dashboard[valor] = formula


        dashboard[valor].font = Font(
            bold=True,
            size=20
        )


        dashboard[valor].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        dashboard[valor].fill = PatternFill("solid", fgColor="FFFFFF")



    dashboard.row_dimensions[3].height = 35
    dashboard.row_dimensions[4].height = 35




    # ==============================
    # TABELA AUXILIAR DO GRÁFICO
    # ==============================


    dashboard["A7"] = "Indicador"
    dashboard["B7"] = "Quantidade"

    for celula in dashboard[7][:2]:
        celula.font = Font(bold=True, color=AZUL_ESCURO)
        celula.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        celula.alignment = Alignment(horizontal="center")


    tabela = [

        (
            "Problemas confirmados",
            '=COUNTIF(\'Dados Analisados\'!G:G,"Problema confirmado")'
        ),


        (
            "Falsos positivos",
            '=COUNTIF(\'Dados Analisados\'!G:G,"Falso positivo")'
        ),


        (
            "Pendentes",
            '=COUNTIF(\'Dados Analisados\'!G:G,"Pendente")'
        )

    ]


    linha = 8


    for nome, formula in tabela:


        dashboard.cell(
            linha,
            1,
            nome
        )


        dashboard.cell(
            linha,
            2,
            formula
        )


        linha +=1




    # ==============================
    # GRÁFICO PIZZA
    # ==============================


    grafico = PieChart()


    grafico.title = "Distribuição das Detecções"



    dados = Reference(

        dashboard,

        min_col=2,

        min_row=7,

        max_row=10

    )


    categorias = Reference(

        dashboard,

        min_col=1,

        min_row=8,

        max_row=10

    )


    grafico.add_data(

        dados,

        titles_from_data=True

    )


    grafico.set_categories(

        categorias

    )


    grafico.height = 12
    grafico.width = 18



    grafico.dataLabels = DataLabelList()

    grafico.dataLabels.showPercent = True

    grafico.dataLabels.showVal = True



    dashboard.add_chart(

        grafico,

        "D7"

    )

    # =====================================
    # GRÁFICO DE CONFUSÕES DA IA
    # =====================================


    ws_dados = wb["Dados Analisados"]


    # Criar tabela auxiliar

    linha_inicio = 15


    dashboard[f"A{linha_inicio}"] = "Confusão IA x Humano"
    dashboard[f"B{linha_inicio}"] = "Quantidade"

    for celula in dashboard[linha_inicio][:2]:
        celula.font = Font(bold=True, color=AZUL_ESCURO)
        celula.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        celula.alignment = Alignment(horizontal="center")


    confusoes = (

        ws_dados

        .values

    )


    from collections import Counter


    contador = Counter()


    for linha in list(confusoes)[1:]:

        ia = linha[2]

        humano = linha[4]


        if (
            humano != "-"
            and humano != "Desconhecido"
        ):

            if (
                ia.lower()
                not in humano.lower()
            ):

                contador[
                    f"{ia} → {humano}"
                ] += 1



    linha = linha_inicio + 1


    for problema, quantidade in contador.items():

        dashboard.cell(
            linha,
            1,
            problema
        )

        dashboard.cell(
            linha,
            2,
            quantidade
        )

        linha += 1



    grafico_barras = BarChart()


    grafico_barras.title = (
        "Principais Confusões da IA"
    )


    grafico_barras.y_axis.title = (
        "Quantidade"
    )


    grafico_barras.x_axis.title = (
        "Erro de classificação"
    )



    dados = Reference(

        dashboard,

        min_col=2,

        min_row=linha_inicio,

        max_row=linha-1

    )


    categorias = Reference(

        dashboard,

        min_col=1,

        min_row=linha_inicio+1,

        max_row=linha-1

    )



    grafico_barras.add_data(

        dados,

        titles_from_data=True

    )



    grafico_barras.set_categories(

        categorias

    )



    grafico_barras.height = 10

    grafico_barras.width = 20



    dashboard.add_chart(

        grafico_barras,

        "K7"

    )



    wb.save(
        caminho
    )

def formatar_analise_erros(caminho):

    wb = load_workbook(caminho)

    if "Análise de Erros IA" not in wb.sheetnames:
        wb.save(caminho)
        return

    ws = wb["Análise de Erros IA"]

    linha_cabecalho = None

    for linha in range(1, ws.max_row + 1):
        if ws.cell(linha, 4).value == "Quantidade":
            linha_cabecalho = linha
            break

    if linha_cabecalho is None or linha_cabecalho == ws.max_row:
        wb.save(caminho)
        return

    azul_escuro = "1F4E78"
    azul_claro = "D9EAF7"
    branco = "FFFFFF"
    borda = Border(
        left=Side(style="thin", color=azul_escuro),
        right=Side(style="thin", color=azul_escuro),
        top=Side(style="thin", color=azul_escuro),
        bottom=Side(style="thin", color=azul_escuro)
    )

    # Título e indicadores são incluídos pela exportação em linhas anteriores
    # ao cabeçalho da tabela.
    if ws["A1"].value == "ANÁLISE DE ERROS DA IA":
        if "A1:D1" not in {str(intervalo) for intervalo in ws.merged_cells.ranges}:
            ws.merge_cells("A1:D1")

        ws["A1"].font = Font(bold=True, size=16, color=branco)
        ws["A1"].fill = PatternFill("solid", fgColor=azul_escuro)
        ws["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        ws.row_dimensions[1].height = 28

        for linha_card in range(3, 6):
            for coluna_card in range(1, 3):
                celula = ws.cell(linha_card, coluna_card)
                celula.fill = PatternFill("solid", fgColor=azul_claro)
                celula.border = borda
                celula.font = Font(bold=True, color=azul_escuro)
                celula.alignment = Alignment(
                    horizontal="center" if coluna_card == 2 else "left",
                    vertical="center"
                )

            ws.row_dimensions[linha_card].height = 22

    for celula in ws[linha_cabecalho]:
        celula.font = Font(bold=True, color=AZUL_ESCURO)
        celula.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        celula.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    ws.auto_filter.ref = (
        f"A{linha_cabecalho}:D{ws.max_row}"
    )
    ws.freeze_panes = f"A{linha_cabecalho + 1}"

    for linha in range(linha_cabecalho + 1, ws.max_row + 1):
        ws.cell(linha, 4).alignment = Alignment(horizontal="center")

    ws.conditional_formatting._cf_rules.clear()
    ws.conditional_formatting.add(
        f"D{linha_cabecalho + 1}:D{ws.max_row}",
        DataBarRule(
            start_type="min",
            end_type="max",
            color="5B9BD5",
            showValue=True
        )
    )

    ajustar_largura_colunas(ws)

    # A aba possui somente este gráfico. Ao reformatar o relatório, remove a
    # versão anterior para evitar gráficos duplicados e referências antigas.
    ws._charts = []

    grafico = BarChart()

    grafico.type = "bar"

    grafico.style = 10

    grafico.title = "Problemas mais confundidos"

    grafico.x_axis.title = "Quantidade"

    grafico.y_axis.title = "Problema"

    ultima_linha = ws.max_row

    dados = Reference(
        ws,
        min_col=4,
        min_row=linha_cabecalho,
        max_row=ultima_linha
    )

    categorias = Reference(
        ws,
        min_col=2,
        min_row=linha_cabecalho + 1,
        max_row=ultima_linha
    )

    grafico.add_data(
        dados,
        titles_from_data=True
    )

    grafico.set_categories(
        categorias
    )

    grafico.height = 10
    grafico.width = 18

    ws.add_chart(
        grafico,
        "F2"
    )

    wb.save(caminho)


def padronizar_tipografia(caminho):

    wb = load_workbook(caminho)

    for ws in wb.worksheets:
        for linha in ws.iter_rows():
            for celula in linha:
                fonte = copy(celula.font)
                fonte.name = FONTE_PADRAO
                celula.font = fonte

    wb.save(caminho)


def formatar_desempenho_ia(caminho):

    wb = load_workbook(caminho)

    if "Desempenho IA" not in wb.sheetnames:
        wb.save(caminho)
        return

    ws = wb["Desempenho IA"]

    linha_cabecalho = None

    for linha in range(1, ws.max_row + 1):
        if ws.cell(linha, 1).value == "IA Detectou":
            linha_cabecalho = linha
            break

    if linha_cabecalho is None:
        wb.save(caminho)
        return

    # Em uma nova execução, remove apenas o painel criado anteriormente. A
    # tabela original permanece na aba e volta a ocupar o início da planilha.
    if ws["A1"].value == "DESEMPENHO DA IA" and linha_cabecalho > 1:
        ws.delete_rows(1, linha_cabecalho - 1)
        linha_cabecalho = 1

    # As barras pertencem ao painel e são recriadas abaixo a cada execução.
    ws.conditional_formatting._cf_rules.clear()

    dados = []

    for linha in range(linha_cabecalho + 1, ws.max_row + 1):
        valores = [ws.cell(linha, coluna).value for coluna in range(1, 8)]

        if any(valor is not None for valor in valores):
            dados.append(valores)

    if not dados:
        wb.save(caminho)
        return

    azul_escuro = "1F4E78"
    azul_claro = "D9EAF7"
    verde = "70AD47"
    vermelho = "C00000"
    amarelo = "FFC000"
    branco = "FFFFFF"
    borda = Border(
        left=Side(style="thin", color="B4C7E7"),
        right=Side(style="thin", color="B4C7E7"),
        top=Side(style="thin", color="B4C7E7"),
        bottom=Side(style="thin", color="B4C7E7")
    )

    altura_painel = 2 + len(dados) * 7
    ws.insert_rows(1, altura_painel)

    ws.merge_cells("A1:E1")
    ws["A1"] = "DESEMPENHO DA IA"
    ws["A1"].font = Font(bold=True, size=16, color=branco)
    ws["A1"].fill = PatternFill("solid", fgColor=azul_escuro)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 26

    nomes_metricas = [
        "Total Detectado",
        "Problemas Confirmados",
        "Falsos Positivos",
        "Pendentes",
        "Precisão (%)"
    ]
    cores_barras = [azul_escuro, verde, vermelho, amarelo, verde]
    maximos_metricas = []

    for coluna in range(1, 5):
        valores = [
            registro[coluna]
            for registro in dados
            if isinstance(registro[coluna], (int, float))
        ]
        maximos_metricas.append(max(valores, default=1))

    linha_inicio = 3

    for indice, registro in enumerate(dados):
        linha_bloco = linha_inicio + indice * 7

        ws.merge_cells(
            start_row=linha_bloco,
            start_column=1,
            end_row=linha_bloco,
            end_column=5
        )
        cabecalho = ws.cell(linha_bloco, 1, str(registro[0]).upper())
        cabecalho.font = Font(bold=True, size=12, color=branco)
        cabecalho.fill = PatternFill("solid", fgColor=azul_escuro)
        cabecalho.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[linha_bloco].height = 23

        valores_metricas = [registro[1], registro[2], registro[3], registro[4]]

        for deslocamento, nome_metrica in enumerate(nomes_metricas):
            linha_metrica = linha_bloco + deslocamento + 1
            ws.merge_cells(
                start_row=linha_metrica,
                start_column=1,
                end_row=linha_metrica,
                end_column=3
            )

            rotulo = ws.cell(linha_metrica, 1, nome_metrica)
            rotulo.fill = PatternFill("solid", fgColor=azul_claro)
            rotulo.font = Font(bold=True, color=azul_escuro)
            rotulo.alignment = Alignment(vertical="center")
            rotulo.border = borda

            valor = ws.cell(linha_metrica, 4)
            valor.border = borda
            valor.font = Font(bold=True, color=azul_escuro)
            valor.alignment = Alignment(horizontal="center", vertical="center")

            if deslocamento < 4:
                valor.value = valores_metricas[deslocamento]
            else:
                valor.value = (
                    f"=IFERROR(D{linha_bloco + 2}/D{linha_bloco + 1},0)"
                )
                valor.number_format = "0.0%"

            barra = ws.cell(linha_metrica, 5)
            barra.value = f"=D{linha_metrica}"
            barra.number_format = ";;;"
            barra.border = borda
            ws.conditional_formatting.add(
                f"E{linha_metrica}",
                DataBarRule(
                    start_type="num",
                    start_value=0,
                    end_type="num",
                    end_value=(
                        1
                        if deslocamento == 4
                        else max(maximos_metricas[deslocamento], 1)
                    ),
                    color=cores_barras[deslocamento],
                    showValue=False
                )
            )

            ws.row_dimensions[linha_metrica].height = 21

    linha_tabela = altura_painel + 1

    for celula in ws[linha_tabela]:
        celula.font = Font(bold=True, color=AZUL_ESCURO)
        celula.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        celula.alignment = Alignment(horizontal="center", vertical="center")

    ws.auto_filter.ref = f"A{linha_tabela}:G{ws.max_row}"
    ws.freeze_panes = f"A{linha_tabela + 1}"

    for coluna in range(2, 8):
        ws.conditional_formatting.add(
            f"{get_column_letter(coluna)}{linha_tabela + 1}:"
            f"{get_column_letter(coluna)}{ws.max_row}",
            DataBarRule(
                start_type="min",
                end_type="max",
                color="5B9BD5",
                showValue=True
            )
        )

    wb.save(caminho)


def formatar_metricas_gerais(caminho):

    wb = load_workbook(caminho)

    if "Métricas Gerais" not in wb.sheetnames:
        wb.save(caminho)
        return

    ws = wb["Métricas Gerais"]

    azul_escuro = "1F4E78"
    cores_cards = {
        "Total capturado": "D9EAF7",
        "Problemas confirmados": "E2F0D9",
        "Falsos positivos": "FCE4D6",
        "Pendentes": "FFF2CC",
        "Classificação correta": "E2F0D9",
        "Classe diferente": "FCE4D6",
        "Classes diferentes": "FCE4D6"
    }
    borda = Border(
        left=Side(style="thin", color="9EADBA"),
        right=Side(style="thin", color="9EADBA"),
        top=Side(style="thin", color="9EADBA"),
        bottom=Side(style="thin", color="9EADBA")
    )

    for celula in ws[1]:
        celula.font = Font(bold=True, color=AZUL_ESCURO, size=11)
        celula.fill = PatternFill("solid", fgColor=AZUL_CLARO)
        celula.alignment = Alignment(horizontal="center", vertical="center")
        celula.border = borda

    ws["C1"] = "Indicador"
    ws["C1"].font = Font(bold=True, color=AZUL_ESCURO, size=11)
    ws["C1"].fill = PatternFill("solid", fgColor=AZUL_CLARO)
    ws["C1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["C1"].border = borda

    ws.row_dimensions[1].height = 24
    ws.conditional_formatting._cf_rules.clear()

    total_capturado = ws.cell(2, 2).value or 0

    try:
        total_capturado = float(total_capturado)
    except (TypeError, ValueError):
        total_capturado = 0

    for linha in range(2, ws.max_row + 1):
        nome_metrica = ws.cell(linha, 1).value
        cor_card = cores_cards.get(nome_metrica, "D9EAF7")

        for coluna in range(1, 3):
            celula = ws.cell(linha, coluna)
            celula.fill = PatternFill("solid", fgColor=cor_card)
            celula.border = borda
            celula.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        ws.cell(linha, 1).font = Font(
            bold=True,
            size=11,
            color=azul_escuro
        )
        ws.cell(linha, 2).font = Font(
            bold=True,
            size=20,
            color=azul_escuro
        )

        valor = ws.cell(linha, 2).value or 0

        try:
            valor = float(valor)
        except (TypeError, ValueError):
            valor = 0

        proporcao = valor / total_capturado if total_capturado else 0

        if nome_metrica == "Problemas confirmados":
            indicador = (
                "🟢 Excelente" if proporcao >= 0.70
                else "🟡 Atenção" if proporcao >= 0.40
                else "🔴 Crítico"
            )
        elif nome_metrica in ("Falsos positivos", "Classe diferente"):
            indicador = (
                "🟢 Excelente" if valor == 0
                else "🟡 Atenção" if proporcao <= 0.10
                else "🔴 Crítico"
            )
        elif nome_metrica == "Pendentes":
            indicador = (
                "🟢 Excelente" if valor == 0
                else "🟡 Atenção" if proporcao <= 0.10
                else "🔴 Crítico"
            )
        else:
            indicador = "—"

        celula_indicador = ws.cell(linha, 3, indicador)
        celula_indicador.fill = PatternFill("solid", fgColor=cor_card)
        celula_indicador.border = borda
        celula_indicador.font = Font(
            name="Segoe UI Emoji",
            bold=True,
            size=11,
            color=azul_escuro
        )
        celula_indicador.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        ws.row_dimensions[linha].height = 38

    ajustar_largura_colunas(ws)
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width, 28)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width, 16)
    ws.column_dimensions["C"].width = 18
    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    ws.freeze_panes = "A2"
    ws.conditional_formatting.add(
        f"B2:B{ws.max_row}",
        DataBarRule(
            start_type="min",
            end_type="max",
            color="5B9BD5",
            showValue=True
        )
    )

    wb.save(caminho)


def criar_resumo_executivo(caminho):

    wb = load_workbook(caminho)

    if "Resumo Executivo" in wb.sheetnames:
        del wb["Resumo Executivo"]

    resumo = wb.create_sheet("Resumo Executivo", 0)

    azul_escuro = AZUL_ESCURO
    azul_claro = AZUL_CLARO
    branco = "FFFFFF"
    borda = Border(
        left=Side(style="thin", color="9EADBA"),
        right=Side(style="thin", color="9EADBA"),
        top=Side(style="thin", color="9EADBA"),
        bottom=Side(style="thin", color="9EADBA")
    )

    for coluna, largura in {
        "A": 25, "B": 14, "C": 4, "D": 25, "E": 14,
        "F": 4, "G": 18, "H": 18, "I": 18, "J": 18,
        "K": 18, "L": 18, "M": 18, "N": 18, "O": 18
    }.items():
        resumo.column_dimensions[coluna].width = largura

    resumo.merge_cells("A1:O1")
    resumo["A1"] = "RESUMO EXECUTIVO DA IA"
    resumo["A1"].font = Font(bold=True, size=18, color=branco)
    resumo["A1"].fill = PatternFill("solid", fgColor=azul_escuro)
    resumo["A1"].alignment = Alignment(horizontal="center", vertical="center")
    resumo.row_dimensions[1].height = 32

    dados = wb["Dados Analisados"]
    cabecalho_dados = [celula.value for celula in dados[1]]

    def indice_coluna(nome):
        return cabecalho_dados.index(nome)

    linhas_dados = list(dados.iter_rows(min_row=2, values_only=True))
    indice_deteccao = indice_coluna("Resultado Detecção")
    indice_classificacao = indice_coluna("Resultado Classificação")
    indice_ia = indice_coluna("IA Detectou")

    total_analisado = len(linhas_dados)
    problemas_confirmados = sum(
        linha[indice_deteccao] == "Problema confirmado"
        for linha in linhas_dados
    )
    falsos_positivos = sum(
        linha[indice_deteccao] == "Falso positivo"
        for linha in linhas_dados
    )
    pendentes = sum(
        linha[indice_deteccao] == "Pendente"
        for linha in linhas_dados
    )
    tipos_detectados = len({
        linha[indice_ia]
        for linha in linhas_dados
        if linha[indice_ia]
    })

    erros = wb["Análise de Erros IA"]
    linha_cabecalho_erros = next(
        linha
        for linha in range(1, erros.max_row + 1)
        if erros.cell(linha, 4).value == "Quantidade"
    )
    erros_ia = [
        {
            "ia": erros.cell(linha, 1).value,
            "objeto": erros.cell(linha, 2).value,
            "familia": erros.cell(linha, 3).value,
            "quantidade": erros.cell(linha, 4).value or 0
        }
        for linha in range(linha_cabecalho_erros + 1, erros.max_row + 1)
        if erros.cell(linha, 2).value is not None
    ]
    erros_ia.sort(key=lambda erro: erro["quantidade"], reverse=True)
    top_cinco = erros_ia[:5]
    total_classes_diferentes = len({
        erro["objeto"] for erro in erros_ia if erro["objeto"]
    })
    total_familias_afetadas = len({
        erro["familia"] for erro in erros_ia if erro["familia"]
    })

    cards = [
        ("A3", "A4:C5", "TOTAL ANALISADO", total_analisado, azul_claro),
        ("D3", "D4:F5", "PROBLEMAS CONFIRMADOS", problemas_confirmados, VERDE),
        ("A7", "A8:C9", "FALSOS POSITIVOS", falsos_positivos, VERMELHO),
        ("D7", "D8:F9", "PENDENTES", pendentes, AMARELO)
    ]

    for titulo, intervalo_valor, nome, valor, cor in cards:
        coluna_inicio = resumo[titulo].column_letter
        linha_titulo = resumo[titulo].row
        coluna_fim = intervalo_valor.split(":")[1][0]
        resumo.merge_cells(f"{titulo}:{coluna_fim}{linha_titulo}")
        resumo.merge_cells(intervalo_valor)

        celula_titulo = resumo[titulo]
        celula_titulo.value = nome
        celula_titulo.font = Font(bold=True, size=10, color=AZUL_ESCURO)
        celula_titulo.fill = PatternFill("solid", fgColor=cor)
        celula_titulo.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        celula_titulo.border = borda

        celula_valor = resumo[intervalo_valor.split(":")[0]]
        celula_valor.value = valor
        celula_valor.font = Font(bold=True, size=22, color=AZUL_ESCURO)
        celula_valor.fill = PatternFill("solid", fgColor=branco)
        celula_valor.alignment = Alignment(horizontal="center", vertical="center")
        celula_valor.border = borda

    resumo.row_dimensions[3].height = 26
    resumo.row_dimensions[4].height = 26
    resumo.row_dimensions[5].height = 26
    resumo.row_dimensions[7].height = 26
    resumo.row_dimensions[8].height = 26
    resumo.row_dimensions[9].height = 26

    resumo.merge_cells("A11:F11")
    resumo["A11"] = "PRINCIPAIS INDICADORES"
    resumo["A11"].font = Font(bold=True, size=12, color=branco)
    resumo["A11"].fill = PatternFill("solid", fgColor=azul_escuro)
    resumo["A11"].alignment = Alignment(horizontal="center")

    indicadores = [
        ("Total de classes diferentes", total_classes_diferentes),
        ("Total de famílias afetadas", total_familias_afetadas),
        ("Tipos de problemas detectados pela IA", tipos_detectados)
    ]

    for linha, (nome, valor) in enumerate(indicadores, start=12):
        resumo.cell(linha, 1, nome)
        resumo.cell(linha, 2, valor)

        for coluna in range(1, 3):
            celula = resumo.cell(linha, coluna)
            celula.fill = PatternFill("solid", fgColor=azul_claro)
            celula.border = borda
            celula.alignment = Alignment(vertical="center")

        resumo.cell(linha, 1).font = Font(bold=True, color=AZUL_ESCURO)
        resumo.cell(linha, 2).font = Font(bold=True, size=14, color=AZUL_ESCURO)
        resumo.cell(linha, 2).alignment = Alignment(horizontal="center")
        resumo.row_dimensions[linha].height = 23

    resumo.merge_cells("A16:F16")
    resumo["A16"] = "TOP 5 CONFUSÕES DA IA"
    resumo["A16"].font = Font(bold=True, size=12, color=branco)
    resumo["A16"].fill = PatternFill("solid", fgColor=azul_escuro)
    resumo["A16"].alignment = Alignment(horizontal="center")

    cabecalhos_top = ["IA Detectou", "Objeto Real", "Família", "Quantidade"]
    for coluna, titulo in enumerate(cabecalhos_top, start=1):
        celula = resumo.cell(17, coluna, titulo)
        celula.font = Font(bold=True, color=AZUL_ESCURO)
        celula.fill = PatternFill("solid", fgColor=azul_claro)
        celula.alignment = Alignment(horizontal="center")
        celula.border = borda

    for linha, erro in enumerate(top_cinco, start=18):
        valores = [
            erro["ia"], erro["objeto"], erro["familia"], erro["quantidade"]
        ]
        for coluna, valor in enumerate(valores, start=1):
            celula = resumo.cell(linha, coluna, valor)
            celula.border = borda
            celula.alignment = Alignment(
                horizontal="center" if coluna == 4 else "left",
                vertical="center",
                wrap_text=True
            )

    if top_cinco:
        resumo.conditional_formatting.add(
            f"D18:D{17 + len(top_cinco)}",
            DataBarRule(
                start_type="min",
                end_type="max",
                color="5B9BD5",
                showValue=True
            )
        )

    principais_classes = [erro["objeto"] for erro in top_cinco[:2]]
    if len(principais_classes) == 2:
        classes_texto = f"{principais_classes[0]} e {principais_classes[1]}"
    elif principais_classes:
        classes_texto = principais_classes[0]
    else:
        classes_texto = "sem classes divergentes registradas"

    resumo.merge_cells("A25:F25")
    resumo["A25"] = "RESUMO AUTOMÁTICO"
    resumo["A25"].font = Font(bold=True, size=12, color=branco)
    resumo["A25"].fill = PatternFill("solid", fgColor=azul_escuro)
    resumo["A25"].alignment = Alignment(horizontal="center")

    resumo.merge_cells("A26:F29")
    resumo["A26"] = (
        f"Foram analisadas {total_analisado} detecções. A IA confirmou "
        f"{problemas_confirmados} problemas, registrou {falsos_positivos} "
        f"falsos positivos e existem {pendentes} pendências. Os erros "
        f"ocorreram principalmente entre as classes {classes_texto}."
    )
    resumo["A26"].font = Font(size=11, color=AZUL_ESCURO)
    resumo["A26"].fill = PatternFill("solid", fgColor="FFFFFF")
    resumo["A26"].border = borda
    resumo["A26"].alignment = Alignment(
        vertical="top", wrap_text=True
    )

    dashboard = wb["Dashboard IA"]

    if dashboard._charts:
        resumo.add_chart(copy(dashboard._charts[0]), "H3")

    if len(dashboard._charts) > 1:
        resumo.add_chart(copy(dashboard._charts[1]), "H22")

    wb.save(caminho)


def refinar_layout_executivo(caminho):

    wb = load_workbook(caminho)

    borda_tabela = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3")
    )

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90
        ws.sheet_properties.tabColor = AZUL_ESCURO

        for linha in range(1, ws.max_row + 1):
            if ws.row_dimensions[linha].height is None:
                ws.row_dimensions[linha].height = 20

    # As tabelas de dados mantêm uma malha discreta e uniforme, sem alterar
    # valores, fórmulas ou filtros já existentes.
    tabelas = {
        "Análise Problemas": (1, 2),
        "Matriz de Erros": (1, 3),
        "Análise de Erros IA": (1, 4),
        "Dados Analisados": (1, 8)
    }

    for nome, (linha_cabecalho, ultima_coluna) in tabelas.items():
        ws = wb[nome]

        for linha in range(linha_cabecalho, ws.max_row + 1):
            for coluna in range(1, ultima_coluna + 1):
                ws.cell(linha, coluna).border = borda_tabela
                ws.cell(linha, coluna).alignment = Alignment(
                    vertical="center",
                    wrap_text=linha == linha_cabecalho
                )

        ws.row_dimensions[linha_cabecalho].height = 24

    dashboard = wb["Dashboard IA"]

    # Os gráficos ficam empilhados à direita: não há mais sobreposição entre
    # eles nem sobre as tabelas auxiliares nas colunas A:B.
    if dashboard._charts:
        dashboard._charts[0].width = 11.5
        dashboard._charts[0].height = 8
        dashboard._charts[0].anchor = "D7"

    if len(dashboard._charts) > 1:
        dashboard._charts[1].width = 11.5
        dashboard._charts[1].height = 8
        dashboard._charts[1].anchor = "D22"

    for linha in range(7, 11):
        for coluna in range(1, 3):
            dashboard.cell(linha, coluna).border = borda_tabela

    for linha in range(15, dashboard.max_row + 1):
        for coluna in range(1, 3):
            dashboard.cell(linha, coluna).border = borda_tabela

    resumo = wb["Resumo Executivo"]

    if resumo._charts:
        resumo._charts[0].width = 12
        resumo._charts[0].height = 8
        resumo._charts[0].anchor = "H3"

    if len(resumo._charts) > 1:
        resumo._charts[1].width = 12
        resumo._charts[1].height = 8
        resumo._charts[1].anchor = "H21"

    # A tabela de desempenho fica separada visualmente do painel de cards.
    desempenho = wb["Desempenho IA"]
    linha_tabela = next(
        linha
        for linha in range(1, desempenho.max_row + 1)
        if desempenho.cell(linha, 1).value == "IA Detectou"
    )

    for linha in range(linha_tabela, desempenho.max_row + 1):
        for coluna in range(1, 8):
            desempenho.cell(linha, coluna).border = borda_tabela

    desempenho.row_dimensions[linha_tabela].height = 24

    wb.save(caminho)


# ==========================================
# EXECUÇÃO
# ==========================================


def gerar_relatorio_visual(caminho):

    formatar_planilhas(caminho)

    criar_dashboard(caminho)

    formatar_analise_erros(caminho)

    formatar_desempenho_ia(caminho)

    formatar_metricas_gerais(caminho)

    criar_resumo_executivo(caminho)

    refinar_layout_executivo(caminho)

    padronizar_tipografia(caminho)






if __name__ == "__main__":


    arquivo = (

        r"C:\Users\marcelino.santos\Desktop"

        r"\Projetos\road-ai-analytics"

        r"\relatorios-gerados"

        r"\Relatorio_Tecnico_IA.xlsx"

    )


    gerar_relatorio_visual(

        arquivo

    )


    print(
        "Relatório formatado com sucesso!"
    )
